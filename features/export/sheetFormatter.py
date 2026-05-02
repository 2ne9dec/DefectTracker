import datetime
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from shared.constants import SEVERITY_LABELS
from shared.utils.dateUtils import fmt_date

class SheetFormatter:
    """Форматирует лист openpyxl с данными дефектов."""

    def fill_sheet(
        self,
        ws,
        rows: list[tuple],
        title_text: str,
        filial: str,
        voltage: str,
        show_fix_cols: bool = False,
    ) -> None:
        fn_bold = Font(name="Times New Roman", size=11, bold=True, color="FFFFFF")
        fn_norm = Font(name="Times New Roman", size=10)
        fn_title = Font(name="Times New Roman", size=14, bold=True, color="FFFFFF")

        al_c = Alignment(horizontal="center", vertical="center", wrap_text=False)
        al_l = Alignment(horizontal="left", vertical="center", wrap_text=True)

        thin = Border(
            left=Side(style="thin", color="BFBFBF"),
            right=Side(style="thin", color="BFBFBF"),
            top=Side(style="thin", color="BFBFBF"),
            bottom=Side(style="thin", color="BFBFBF"),
        )

        fill_hdr = PatternFill(start_color="1F3864", fill_type="solid")
        fill_alt = PatternFill(start_color="EFF3FF", fill_type="solid")
        fill_wht = PatternFill(start_color="FFFFFF", fill_type="solid")
        sev_fills = {
            "critical": PatternFill(start_color="FFD7D7", fill_type="solid"),
            "medium": PatternFill(start_color="FFF0CC", fill_type="solid"),
            "low": PatternFill(start_color="D7F5DE", fill_type="solid"),
        }

        num_cols = 9 if show_fix_cols else 7
        last_col = get_column_letter(num_cols)

        # Заголовок листа
        ws.merge_cells(f"A1:{last_col}1")
        c = ws["A1"]
        c.value = title_text
        c.font = fn_title
        c.alignment = al_c
        c.fill = PatternFill(start_color="1F3864", fill_type="solid")
        ws.row_dimensions[1].height = 32

        # Мета-информация
        ws.merge_cells(f"A3:{last_col}3")
        ws["A3"].value = (
            f"Филиал: {filial}    |    "
            f"Напряжение: {voltage}    |    "
            f"Дата выгрузки: {datetime.date.today():%d.%m.%Y}"
        )
        ws["A3"].font = Font(name="Times New Roman", size=10, italic=True)

        # Шапка таблицы
        if show_fix_cols:
            headers = [
                "Опора",
                "Элемент",
                "Дефект",
                "Серьёзность",
                "Дата обн.",
                "Обнаружил",
                "Дата устр.",
                "Устранил",
                "Статус",
            ]
        else:
            headers = [
                "Опора",
                "Элемент",
                "Дефект",
                "Серьёзность",
                "Дата обн.",
                "Обнаружил",
                "Статус",
            ]

        HDR_ROW = 5
        ws.row_dimensions[4].height = 5
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=HDR_ROW, column=ci, value=h)
            cell.font = fn_bold
            cell.fill = fill_hdr
            cell.border = thin
            cell.alignment = al_c

        # Строки данных
        total_row_idx = HDR_ROW + 1
        for ri, row in enumerate(rows, 1):
            er = HDR_ROW + ri
            pole, elem, defect, sev, date_f, insp_f, date_fix, insp_fix, status = row
            row_data = [
                pole,
                elem,
                defect,
                SEVERITY_LABELS.get(sev, sev),
                fmt_date(date_f),
                insp_f,
            ]
            if show_fix_cols:
                row_data += [fmt_date(date_fix) if date_fix else "-", insp_fix or "-"]
            row_data.append(status)

            row_fill = sev_fills.get(sev, fill_alt if ri % 2 == 0 else fill_wht)
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=er, column=ci, value=val)
                c.font = fn_norm
                c.border = thin
                c.fill = row_fill
                c.alignment = al_c if ci in [1, 5, 7, num_cols] else al_l
            total_row_idx = er

        # Итоговая строка и автофильтр
        if rows:
            ws.auto_filter.ref = f"A{HDR_ROW}:{last_col}{total_row_idx}"
            sum_row = total_row_idx + 2
            ws.cell(row=sum_row, column=1, value="ИТОГО дефектов:").font = Font(
                name="Times New Roman", size=11, bold=True
            )
            ws.cell(row=sum_row, column=2, value=len(rows)).font = Font(
                name="Times New Roman", size=11, bold=True
            )

        # Ширина колонок
        col_widths = (
            [10, 28, 42, 14, 14, 22, 14, 22, 12]
            if show_fix_cols
            else [10, 28, 42, 14, 14, 22, 12]
        )
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
